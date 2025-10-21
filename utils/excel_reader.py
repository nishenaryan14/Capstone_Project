import openpyxl, os
from typing import List, Dict

def read_excel(file_path: str, sheet: str = None) -> List[Dict]:
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet] if sheet else wb.active
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data.append(dict(zip(headers, row)))
    wb.close()
    return data